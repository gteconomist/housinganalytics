#!/usr/bin/env python3
"""
build-data.py
─────────────
Mirror of scripts/build-data.mjs in Python, used for local sandboxed builds
(and as a manual fallback if npm is unavailable). Reads
`Full Housing Data Table.xlsx` and emits per-county JSON to src/data/generated/.

The canonical build path is the Node script; this file exists so the data
layer can be exercised without npm.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl  # pip install openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / 'Full Housing Data Table.xlsx'
OUT_DIR = ROOT / 'src' / 'data' / 'generated'


# Mirror of src/data/variable-map.js
VARIABLES = [
    # demographics
    ('population_total',          'B01003_001E'),
    ('median_age',                'DP05_0018E'),
    ('race_white',                'DP05_0037E'),
    ('race_black',                'DP05_0045E'),
    ('race_aian',                 'DP05_0053E'),
    ('race_asian',                'DP05_0061E'),
    ('race_nhpi',                 'DP05_0069E'),
    ('race_other',                'DP05_0074E'),
    ('race_two_plus',             'DP05_0075E'),
    # occupancy & structure
    ('units_occupied',            'B25002_002E'),
    ('units_vacant',              'B25002_003E'),
    ('units_total',               'B25024_001E'),
    ('structure_1_detached',      'B25024_002E'),
    ('structure_1_attached',      'B25024_003E'),
    ('structure_2',               'B25024_004E'),
    ('structure_3_4',             'B25024_005E'),
    ('structure_5_9',             'B25024_006E'),
    ('structure_10_19',           'B25024_007E'),
    ('structure_20_49',           'B25024_008E'),
    ('structure_50_plus',         'B25024_009E'),
    ('structure_mobile',          'B25024_010E'),
    ('structure_other',           'B25024_011E'),
    ('structure_median_age',      'B25035_001E'),
    # value
    ('value_median',              'DP04_0089E'),
    ('value_lt_50k',              'DP04_0081E'),
    ('value_50_99k',              'DP04_0082E'),
    ('value_100_150k',            'DP04_0083E'),
    ('value_150_200k',            'DP04_0084E'),
    ('value_200_300k',            'DP04_0085E'),
    ('value_300_500k',            'DP04_0086E'),
    ('value_500_1m',              'DP04_0087E'),
    ('value_1m_plus',             'DP04_0088E'),
    # tenure
    ('tenure_total_occupied',     'DP04_0045E'),
    ('tenure_owner_occupied',     'DP04_0046E'),
    ('tenure_renter_occupied',    'DP04_0047E'),
    # income
    ('poverty_rate',              'S1701_C03_006E'),
    ('hh_income_median',          'S1901_C01_012E'),
    ('hh_income_mean',            'S1901_C01_013E'),
    ('per_capita_income',         'DP03_0088E'),
    # year built
    ('year_built_2020_plus',      'S2504_C01_009E'),
    ('year_built_2010_19',        'S2504_C01_010E'),
    ('year_built_2000_09',        'S2504_C01_011E'),
    ('year_built_1980_99',        'S2504_C01_012E'),
    ('year_built_1960_79',        'S2504_C01_013E'),
    ('year_built_1940_59',        'S2504_C01_014E'),
    ('year_built_1939_earlier',   'S2504_C01_015E'),
    # earners / household
    ('earners_0',                 'S1903_C01_030E'),
    ('earners_1',                 'S1903_C01_031E'),
    ('earners_2',                 'S1903_C01_032E'),
    ('earners_3_plus',            'S1903_C01_033E'),
    ('earners_nonfamily',         'S1903_C01_034E'),
    ('earners_0_income',          'S1903_C03_030E'),
    ('earners_1_income',          'S1903_C03_031E'),
    ('earners_2_income',          'S1903_C03_032E'),
    ('earners_3_plus_income',     'S1903_C03_033E'),
    ('earners_nonfamily_income',  'S1903_C03_034E'),
    ('hh_total_s1903',            'S1903_C01_001E'),
    ('hh_size_1',                 'S2501_C01_002E'),
    ('hh_size_2',                 'S2501_C01_003E'),
    ('hh_size_3',                 'S2501_C01_004E'),
    ('hh_size_4_plus',            'S2501_C01_005E'),
    ('hh_total_s2501',            'S2501_C01_001E'),
    ('hh_total_s1101',            'S1101_C01_001E'),
    ('hh_avg_size',               'S1101_C01_002E'),
    ('hh_with_children',          'S1101_C01_005E'),
    # mortgage
    ('mortgage_total',            'S2506_C01_001E'),
    ('mortgage_median',           'S2506_C01_009E'),
    # education
    ('edu_lt_9th',                'S1501_C01_007E'),
    ('edu_9_12_no_diploma',       'S1501_C01_008E'),
    ('edu_hs_grad',               'S1501_C01_009E'),
    ('edu_some_college',          'S1501_C01_010E'),
    ('edu_associates',            'S1501_C01_011E'),
    ('edu_bachelors',             'S1501_C01_012E'),
    ('edu_graduate',              'S1501_C01_013E'),
    # MHC coarse buckets (1000-1499..3000+)
    ('mhc_1000_1499',             'B25104_012E'),
    ('mhc_1500_1999',             'B25104_013E'),
    ('mhc_2000_2499',             'B25104_014E'),
    ('mhc_2500_2999',             'B25104_015E'),
    ('mhc_3000_plus',             'B25104_016E'),
    ('mhc_no_cash_rent',          'B25104_017E'),
    # rent
    ('rent_median',               'B25031_001E'),
    ('rent_median_0br',           'B25031_002E'),
    ('rent_median_1br',           'B25031_003E'),
    ('rent_median_2br',           'B25031_004E'),
    ('rent_median_3br',           'B25031_005E'),
    ('rent_median_4br',           'B25031_006E'),
    ('rent_median_5br',           'B25031_007E'),
    # rent burden (B25070)
    ('rent_burden_total',         'B25070_001E'),
    ('rent_burden_lt_10',         'B25070_002E'),
    ('rent_burden_10_14',         'B25070_003E'),
    ('rent_burden_15_19',         'B25070_004E'),
    ('rent_burden_20_24',         'B25070_005E'),
    ('rent_burden_25_29',         'B25070_006E'),
    ('rent_burden_30_34',         'B25070_007E'),
    ('rent_burden_35_39',         'B25070_008E'),
    ('rent_burden_40_49',         'B25070_009E'),
    ('rent_burden_50_plus',       'B25070_010E'),
    ('rent_burden_not_computed',  'B25070_011E'),
    # owner cost burden (B25106)
    ('cb_owner_total',            'B25106_002E'),
    ('cb_owner_30_plus_lt20k',    'B25106_006E'),
    ('cb_owner_30_plus_20_35k',   'B25106_010E'),
    ('cb_owner_30_plus_35_50k',   'B25106_014E'),
    ('cb_owner_30_plus_50_75k',   'B25106_018E'),
    ('cb_owner_30_plus_75k_plus', 'B25106_022E'),
    ('cb_owner_zero_negative',    'B25106_023E'),
    ('cb_renter_total',           'B25106_024E'),
    ('cb_renter_no_cash',         'B25106_046E'),
    # bedrooms
    ('br_0',  'B25041_002E'),
    ('br_1',  'B25041_003E'),
    ('br_2',  'B25041_004E'),
    ('br_3',  'B25041_005E'),
    ('br_4',  'B25041_006E'),
    ('br_5_plus', 'B25041_007E'),
    # owner/renter income (B25118)
    ('oi_lt_5k',     'B25118_003E'),
    ('oi_5_10k',     'B25118_004E'),
    ('oi_10_15k',    'B25118_005E'),
    ('oi_15_20k',    'B25118_006E'),
    ('oi_20_25k',    'B25118_007E'),
    ('oi_25_35k',    'B25118_008E'),
    ('oi_35_50k',    'B25118_009E'),
    ('oi_50_75k',    'B25118_010E'),
    ('oi_75_100k',   'B25118_011E'),
    ('oi_100_150k',  'B25118_012E'),
    ('oi_150k_plus', 'B25118_013E'),
    ('ri_lt_5k',     'B25118_015E'),
    ('ri_5_10k',     'B25118_016E'),
    ('ri_10_15k',    'B25118_017E'),
    ('ri_15_20k',    'B25118_018E'),
    ('ri_20_25k',    'B25118_019E'),
    ('ri_25_35k',    'B25118_020E'),
    ('ri_35_50k',    'B25118_021E'),
    ('ri_50_75k',    'B25118_022E'),
    ('ri_75_100k',   'B25118_023E'),
    ('ri_100_150k',  'B25118_024E'),
    ('ri_150k_plus', 'B25118_025E'),
    # tenure × age (S2502)
    ('renter_age_under35', 'S2502_C05_011E'),
    ('renter_age_35_44',   'S2502_C05_012E'),
    ('renter_age_45_54',   'S2502_C05_013E'),
    ('renter_age_55_64',   'S2502_C05_014E'),
    ('renter_age_65_74',   'S2502_C05_015E'),
    ('renter_age_75_84',   'S2502_C05_016E'),
    ('renter_age_85_plus', 'S2502_C05_017E'),
    ('owner_age_under35',  'S2502_C03_011E'),
    ('owner_age_35_44',    'S2502_C03_012E'),
    ('owner_age_45_54',    'S2502_C03_013E'),
    ('owner_age_55_64',    'S2502_C03_014E'),
    ('owner_age_65_74',    'S2502_C03_015E'),
    ('owner_age_75_84',    'S2502_C03_016E'),
    ('owner_age_85_plus',  'S2502_C03_017E'),
]

ACS_TO_FIELD = {acs: field for field, acs in VARIABLES}

MHC_LT_500_CODES   = ['B25104_002E', 'B25104_003E', 'B25104_004E', 'B25104_005E', 'B25104_006E']
MHC_500_999_CODES  = ['B25104_007E', 'B25104_008E', 'B25104_009E', 'B25104_010E', 'B25104_011E']

SUM_FIELDS = {
    'population_total', 'units_occupied', 'units_vacant', 'units_total',
    'structure_1_detached', 'structure_1_attached', 'structure_2', 'structure_3_4',
    'structure_5_9', 'structure_10_19', 'structure_20_49', 'structure_50_plus',
    'structure_mobile', 'structure_other',
    'tenure_total_occupied', 'tenure_owner_occupied', 'tenure_renter_occupied',
    'year_built_2020_plus', 'year_built_2010_19', 'year_built_2000_09', 'year_built_1980_99',
    'year_built_1960_79', 'year_built_1940_59', 'year_built_1939_earlier',
    'earners_0', 'earners_1', 'earners_2', 'earners_3_plus', 'earners_nonfamily',
    'hh_size_1', 'hh_size_2', 'hh_size_3', 'hh_size_4_plus',
    'hh_total_s2501', 'hh_total_s1101', 'hh_total_s1903', 'hh_with_children',
    'br_0', 'br_1', 'br_2', 'br_3', 'br_4', 'br_5_plus',
    'rent_burden_total', 'rent_burden_lt_10', 'rent_burden_10_14', 'rent_burden_15_19',
    'rent_burden_20_24', 'rent_burden_25_29', 'rent_burden_30_34', 'rent_burden_35_39',
    'rent_burden_40_49', 'rent_burden_50_plus', 'rent_burden_not_computed',
    'cb_owner_total', 'cb_owner_30_plus_lt20k', 'cb_owner_30_plus_20_35k',
    'cb_owner_30_plus_35_50k', 'cb_owner_30_plus_50_75k', 'cb_owner_30_plus_75k_plus',
    'cb_owner_zero_negative', 'cb_renter_total', 'cb_renter_no_cash',
    'edu_lt_9th', 'edu_9_12_no_diploma', 'edu_hs_grad', 'edu_some_college',
    'edu_associates', 'edu_bachelors', 'edu_graduate',
    'value_lt_50k', 'value_50_99k', 'value_100_150k', 'value_150_200k',
    'value_200_300k', 'value_300_500k', 'value_500_1m', 'value_1m_plus',
    'oi_lt_5k', 'oi_5_10k', 'oi_10_15k', 'oi_15_20k', 'oi_20_25k', 'oi_25_35k',
    'oi_35_50k', 'oi_50_75k', 'oi_75_100k', 'oi_100_150k', 'oi_150k_plus',
    'ri_lt_5k', 'ri_5_10k', 'ri_10_15k', 'ri_15_20k', 'ri_20_25k', 'ri_25_35k',
    'ri_35_50k', 'ri_50_75k', 'ri_75_100k', 'ri_100_150k', 'ri_150k_plus',
    'renter_age_under35', 'renter_age_35_44', 'renter_age_45_54', 'renter_age_55_64',
    'renter_age_65_74', 'renter_age_75_84', 'renter_age_85_plus',
    'owner_age_under35', 'owner_age_35_44', 'owner_age_45_54', 'owner_age_55_64',
    'owner_age_65_74', 'owner_age_75_84', 'owner_age_85_plus',
    'mhc_lt_500', 'mhc_500_999', 'mhc_1000_1499', 'mhc_1500_1999', 'mhc_2000_2499',
    'mhc_2500_2999', 'mhc_3000_plus', 'mhc_no_cash_rent',
    'mortgage_total',
    'race_white', 'race_black', 'race_aian', 'race_asian', 'race_nhpi',
    'race_other', 'race_two_plus',
}

WEIGHTED_MEDIAN_FIELDS = [
    ('hh_income_median',  'tenure_total_occupied'),
    ('hh_income_mean',    'tenure_total_occupied'),
    ('value_median',      'tenure_owner_occupied'),
    ('rent_median',       'tenure_renter_occupied'),
    ('rent_median_0br',   'br_0'),
    ('rent_median_1br',   'br_1'),
    ('rent_median_2br',   'br_2'),
    ('rent_median_3br',   'br_3'),
    ('rent_median_4br',   'br_4'),
    ('rent_median_5br',   'br_5_plus'),
    ('mortgage_median',   'mortgage_total'),
    ('structure_median_age', 'units_total'),
    ('median_age',        'population_total'),
    ('per_capita_income', 'population_total'),
    ('hh_avg_size',       'tenure_total_occupied'),
    ('poverty_rate',      'population_total'),
]

def num(v):
    if v is None or v == '':
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    # ACS sentinel values for "no estimate available" / "can't be computed":
    #   -666666666, -888888888, -999999999. Treat as null.
    if n < -1_000_000:
        return None
    return n

def safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b

def pct(a, b):
    v = safe_div(a, b)
    return None if v is None else v * 100

def slug(s):
    s = (s or '').lower()
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def add_derived(c):
    c['homeownership_rate'] = pct(c.get('tenure_owner_occupied'), c.get('tenure_total_occupied'))
    c['renter_rate']        = pct(c.get('tenure_renter_occupied'), c.get('tenure_total_occupied'))
    total_units = (c.get('units_occupied') or 0) + (c.get('units_vacant') or 0)
    c['vacancy_rate'] = pct(c.get('units_vacant'), total_units or None)

    rb_30 = sum(c.get(k) or 0 for k in
                ['rent_burden_30_34', 'rent_burden_35_39', 'rent_burden_40_49', 'rent_burden_50_plus'])
    rb_severe = c.get('rent_burden_50_plus') or 0
    rb_denom = (c.get('rent_burden_total') or 0) - (c.get('rent_burden_not_computed') or 0)
    c['renter_cost_burden_rate']        = pct(rb_30, rb_denom or None)
    c['renter_severe_cost_burden_rate'] = pct(rb_severe, rb_denom or None)

    ob_30 = sum(c.get(k) or 0 for k in
                ['cb_owner_30_plus_lt20k', 'cb_owner_30_plus_20_35k', 'cb_owner_30_plus_35_50k',
                 'cb_owner_30_plus_50_75k', 'cb_owner_30_plus_75k_plus'])
    c['owner_cost_burden_rate'] = pct(ob_30, c.get('cb_owner_total'))

    c['price_to_income_ratio'] = safe_div(c.get('value_median'), c.get('hh_income_median'))
    c['rent_to_income_ratio'] = (
        (c['rent_median'] * 12) / c['hh_income_median']
        if c.get('rent_median') is not None and c.get('hh_income_median')
        else None
    )

    yb_total = sum(c.get(k) or 0 for k in
                   ['year_built_2020_plus', 'year_built_2010_19', 'year_built_2000_09',
                    'year_built_1980_99', 'year_built_1960_79', 'year_built_1940_59',
                    'year_built_1939_earlier'])
    pre_1980 = sum(c.get(k) or 0 for k in
                   ['year_built_1960_79', 'year_built_1940_59', 'year_built_1939_earlier'])
    c['aging_stock_share'] = pct(pre_1980, yb_total or None)

    sf = (c.get('structure_1_detached') or 0) + (c.get('structure_1_attached') or 0)
    c['single_family_share'] = pct(sf, c.get('units_total'))

    mm = sum(c.get(k) or 0 for k in
             ['structure_2', 'structure_3_4', 'structure_5_9', 'structure_10_19'])
    c['missing_middle_share'] = pct(mm, c.get('units_total'))

    edu_total = sum(c.get(k) or 0 for k in
                    ['edu_lt_9th', 'edu_9_12_no_diploma', 'edu_hs_grad', 'edu_some_college',
                     'edu_associates', 'edu_bachelors', 'edu_graduate'])
    c['bachelors_plus_rate'] = pct((c.get('edu_bachelors') or 0) + (c.get('edu_graduate') or 0),
                                   edu_total or None)
    return c


def main():
    print(f'Reading {SOURCE_XLSX}...')
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))

    header = rows[11]
    col_to_field = {}
    for col, val in enumerate(header):
        if val is None:
            continue
        s = str(val).strip()
        if s in ACS_TO_FIELD:
            col_to_field[col] = ACS_TO_FIELD[s]

    try:
        NAME_COL  = header.index('NAME')
        GEOID_COL = header.index('Geo_geoid_')
    except ValueError:
        print('ERROR: missing NAME or Geo_geoid_ column on Excel row 12', file=sys.stderr)
        sys.exit(1)

    mhc_lt_500_cols  = [header.index(c) for c in MHC_LT_500_CODES  if c in header]
    mhc_500_999_cols = [header.index(c) for c in MHC_500_999_CODES if c in header]

    counties = []
    for row in rows[12:]:
        if not row:
            continue
        name = row[NAME_COL]
        geo = row[GEOID_COL]
        if not name or not geo:
            continue
        geoid = str(geo).strip()
        if geoid.isdigit():
            geoid = geoid.zfill(5)
        county_name, _, state_name = (str(name) + ',').partition(',')
        state_name = state_name.strip().rstrip(',').strip()

        c = {
            'geoid': geoid,
            'name': name,
            'county_name': county_name.strip(),
            'state_name': state_name,
            'state_fips': geoid[:2],
            'slug': slug(county_name),
            'state_slug': slug(state_name),
        }
        for col, field in col_to_field.items():
            c[field] = num(row[col])

        # MHC derived buckets
        def sum_cols(cols):
            vals = [num(row[col]) for col in cols]
            kept = [v for v in vals if v is not None]
            return sum(kept) if kept else None
        c['mhc_lt_500']  = sum_cols(mhc_lt_500_cols)
        c['mhc_500_999'] = sum_cols(mhc_500_999_cols)

        add_derived(c)
        counties.append(c)

    print(f'Parsed {len(counties)} counties.')

    # Aggregates
    def aggregate(group, name, type_):
        out = {'name': name, 'type': type_}
        for f in SUM_FIELDS:
            s = 0
            any_ = False
            for c in group:
                v = c.get(f)
                if v is not None:
                    s += v
                    any_ = True
            out[f] = s if any_ else None
        for field, weight in WEIGHTED_MEDIAN_FIELDS:
            num_, denom = 0.0, 0.0
            for c in group:
                v = c.get(field)
                w = c.get(weight)
                if v is not None and w is not None and w > 0:
                    num_  += v * w
                    denom += w
            out[field] = (num_ / denom) if denom > 0 else None
        add_derived(out)
        out['county_count'] = len(group)
        return out

    national = aggregate(counties, 'United States', 'national')

    by_state = {}
    for c in counties:
        by_state.setdefault(c['state_name'], []).append(c)
    state_aggs = {}
    for state, group in by_state.items():
        agg = aggregate(group, state, 'state')
        agg['state_fips'] = group[0]['state_fips'] if group else None
        state_aggs[state] = agg

    def light_context(a):
        return {k: a.get(k) for k in [
            'name', 'population_total', 'hh_income_median', 'value_median', 'rent_median',
            'homeownership_rate', 'renter_rate', 'vacancy_rate',
            'renter_cost_burden_rate', 'owner_cost_burden_rate',
            'price_to_income_ratio', 'aging_stock_share',
            'single_family_share', 'bachelors_plus_rate',
        ]}

    # Write outputs
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / 'counties').mkdir(exist_ok=True)
    (OUT_DIR / 'states').mkdir(exist_ok=True)

    for c in counties:
        state_agg = state_aggs.get(c['state_name'])
        c['_context'] = {
            'state':    light_context(state_agg) if state_agg else None,
            'national': light_context(national),
        }
        (OUT_DIR / 'counties' / f'{c["geoid"]}.json').write_text(json.dumps(c, indent=2))

    for state, agg in state_aggs.items():
        (OUT_DIR / 'states' / f'{slug(state)}.json').write_text(json.dumps(agg, indent=2))

    (OUT_DIR / 'national.json').write_text(json.dumps(national, indent=2))

    manifest = {
        'generated_at': __import__('datetime').datetime.utcnow().isoformat() + 'Z',
        'source_file': 'Full Housing Data Table.xlsx',
        'county_count': len(counties),
        'state_count': len(state_aggs),
        'states': sorted(
            ({
                'name': s['name'],
                'slug': slug(s['name']),
                'state_fips': s.get('state_fips'),
                'county_count': s['county_count'],
                'population_total': s.get('population_total'),
            } for s in state_aggs.values()),
            key=lambda s: s['name'],
        ),
        'counties': sorted(
            ({
                'geoid': c['geoid'],
                'name': c['county_name'],
                'state': c['state_name'],
                'state_slug': c['state_slug'],
                'slug': c['slug'],
                'population_total': c.get('population_total'),
                'hh_income_median': c.get('hh_income_median'),
                'value_median': c.get('value_median'),
                'homeownership_rate': c.get('homeownership_rate'),
            } for c in counties),
            key=lambda c: (c['state'], c['name']),
        ),
    }
    (OUT_DIR / 'manifest.json').write_text(json.dumps(manifest, indent=2))

    print(f'✓ Wrote {len(counties)} counties, {len(state_aggs)} states, 1 national to {OUT_DIR}')


if __name__ == '__main__':
    main()
